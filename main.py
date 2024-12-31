import gradio as gr
from setup import *
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from agents import research_agent
from vectorstore import extract_urls, urls_classify_list, clean_and_extract_html_data
from usecase_agent import usecase_agent_func, vectorstore_writing
# from feasibility_agent import feasibility_agent_func



# # Function to create Excel file
# def create_excel(df):
#     # Create a new Excel workbook and select the active sheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Use Cases"

#     # Define and write headers to the Excel sheet
#     headers = ['Use Case', 'Description', 'URLs']
#     ws.append(headers)

#     # Write data rows
#     for _, row in df.iterrows():
#         try:
#             use_case = row['use_case']
#             description = row['description']
#             urls = row['urls_list']

#             ws.append([use_case, description, None])  # Add use case and description
#             if urls:
#                 for url_index, url in enumerate(urls):
#                     cell = ws.cell(row=ws.max_row, column=3)  # URLs go into the third column
#                     cell.value = url
#                     cell.hyperlink = url
#                     cell.font = Font(color="0000FF", underline="single")
                    
#                     # Add a new row for additional URLs
#                     if url_index < len(urls) - 1:
#                         ws.append([None, None, None])
#         except KeyError as e:
#             print(f"Missing key in DataFrame row: {e}")
#         except Exception as e:
#             print(f"Unexpected error while processing row: {e}")

#     excel_file_path = "GenAI_use_cases_feasibility.xlsx"
#     wb.save(excel_file_path)
#     return excel_file_path


# # Function to handle the report and create the DataFrame
# def pd_creation(report):
#     # Assuming feasibility_agent_func returns a dictionary
#     pd_dict = feasibility_agent_func(report)
    
#     # Check for expected keys in pd_dict before proceeding
#     required_columns = ['use_case', 'description', 'urls_list']
#     if not all(col in pd_dict for col in required_columns):
#         raise ValueError(f"Missing one or more expected columns: {required_columns}")

#     # Create the DataFrame from the dictionary
#     df = pd.DataFrame(pd_dict)
    
#     # Convert the dataframe to the format expected by Gradio (list of lists)
#     data = df.values.tolist()  # This creates a list of lists from the dataframe
    
#     # Create the Excel file and return its path
#     excel_file_path = create_excel(df)  # Create the Excel file and get its path
    
#     return data, excel_file_path  # Return the formatted data and the Excel file path

# Main function that handles the user query and generates the report
def main(user_input):
    # Research Agent
    agentstate_result = research_agent(user_input)

    # Vector Store
    urls, content = extract_urls(agentstate_result)
    pdf_urls, html_urls = urls_classify_list(urls)
    html_docs = clean_and_extract_html_data(html_urls)

    # Writing vector store (not explicitly defined in your example)
    vectorstore_writing(html_docs)

    # Use-case agent
    company_name = agentstate_result['company']
    industry_name = agentstate_result['industry']

    if company_name:
        topic = f'GenAI Usecases in {company_name} and {industry_name} industry. Explore {company_name} GenAI applications, key offerings, strategic focus areas, competitors, and market share.'
    else:
        topic = f'GenAI Usecases in {industry_name}. Explore {industry_name} GenAI applications, trends, challenges, and opportunities.'
    max_analysts = 3

    report = usecase_agent_func(topic, max_analysts)
    # pd_dict, excel_file_path = pd_creation(report)
    
    # Save the report as a markdown file
    report_file_path = "generated_report.md"
    with open(report_file_path, "w") as f:
        f.write(report)
    # pd_dict, excel_file_path
    return report, report_file_path 

# Example queries
examples = [
    "How is the retail industry leveraging AI and ML?",
    "AI applications in automotive manufacturing"
]

# Creating the Gradio interface
with gr.Blocks(theme=gr.themes.Soft(font=gr.themes.GoogleFont('Open Sans'))) as demo:
    # Header section
    gr.HTML("<center><h1>UseCaseGenie - Discover GenAI Use cases for your company and Industry! 🤖🧑‍🍳.</h1><center>") 
    gr.Markdown("""#### This GenAI Assistant 🤖 helps you discover and explore Generative AI use cases for your company and industry. 
    You can download the generated use case report as a <b>Markdown file</b> to gain insights and explore relevant GenAI applications.
    ### <b>Steps:</b>
    1. <b>Enter your query</b> regarding any company or industry.  
    2. <b>Click on the 'Submit' button</b> and wait for the GenAI assistant to generate the report.  
    3. <b>Download the generated report<b> 
    4. Explore the GenAI use cases and URLs for further analysis. 
    """)


    # Input for the user query
    with gr.Row():
        user_input = gr.Textbox(label="Enter your Query", placeholder='Type_here...')

    # Examples to help users with inputs
    with gr.Row():  
        gr.Examples(examples=examples, inputs=user_input)

    # Buttons for submitting and downloading
    with gr.Row():
        submit_button = gr.Button("Submit")
        clear_btn = gr.ClearButton([user_input], value='Clear')

    # File download buttons
    with gr.Row():
        # Create a downloadable markdown file
        download_report_button = gr.File(label="Usecases Report")    

        # # Create a downloadable Excel file
        # download_excel_button = gr.File(label="Feasibility Excel File")   

    # Display report in Markdown format
    with gr.Row():
        report_output = gr.Markdown()

    submit_button.click(main, inputs=[user_input], outputs=[report_output, download_report_button])

# Run the interface
demo.launch()
